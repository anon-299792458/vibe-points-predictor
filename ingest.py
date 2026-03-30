#!/usr/bin/env python3
"""
Data ingestion for CAO 2026 Point Predictor.
Parses CAO cut-off PDFs/xlsx (2017-2025) and LC points distribution PDFs
into a local SQLite database (predictor.db).

Run once (or re-run to rebuild):
    python ingest.py
"""

import re
import sys
import sqlite3
from pathlib import Path

import pdfplumber
import openpyxl
from rapidfuzz import fuzz

DATA_DIR = Path(__file__).parent / "Data"
DB_PATH  = Path(__file__).parent / "predictor.db"

# Grade adjustment factors confirmed by Dept of Education (Nov 2025).
# 2017–2019: baseline (no post-marking uplift)
# 2020–2024: held at ~2020 "Calculated Grades" plateau (factor 1.0)
# 2025:      partial removal — midway between 2020 and 2019 levels (factor 0.6)
GRADE_ADJ: dict[int, float] = {
    2017: 0.0, 2018: 0.0, 2019: 0.0,
    2020: 1.0, 2021: 1.0, 2022: 1.0, 2023: 1.0, 2024: 1.0,
    2025: 0.6,
}
# is_anomaly = True for years with heavy grade inflation (used for chart colouring)
ANOMALY_YEARS = {y for y, f in GRADE_ADJ.items() if f >= 0.5}  # 2020-2024
RECENT_YEARS  = {2023, 2024, 2025}


# ─────────────────────────────────────────────────────────────────────────────
# Database
# ─────────────────────────────────────────────────────────────────────────────

DDL = """
DROP TABLE IF EXISTS course_points;
DROP TABLE IF EXISTS cohort_stats;

CREATE TABLE course_points (
    id                      INTEGER PRIMARY KEY AUTOINCREMENT,
    year                    INTEGER NOT NULL,
    institution             TEXT,
    course_code             TEXT    NOT NULL,
    course_name             TEXT    NOT NULL,
    points                  INTEGER,
    is_capped               INTEGER NOT NULL DEFAULT 0,
    is_anomaly              INTEGER NOT NULL DEFAULT 0,
    grade_adjustment_factor REAL    NOT NULL DEFAULT 0.0,
    match_confidence        REAL    DEFAULT 1.0,
    source                  TEXT
);

CREATE TABLE cohort_stats (
    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
    year               INTEGER NOT NULL,
    points_band_lower  INTEGER NOT NULL,
    points_band_upper  INTEGER NOT NULL,
    student_count      INTEGER NOT NULL,
    cumulative_pct     REAL
);

CREATE INDEX idx_cp_code ON course_points(course_code);
CREATE INDEX idx_cp_year ON course_points(year);
CREATE INDEX idx_cs_year ON cohort_stats(year);
"""


def create_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(DDL)
    conn.commit()
    return conn


# ─────────────────────────────────────────────────────────────────────────────
# Shared utilities
# ─────────────────────────────────────────────────────────────────────────────

def to_int_points(val):
    """
    Convert a raw cell / string value to int points (0-625), or None.
    Strips leading #, *, + and trailing *, +.
    Returns None if the value is non-numeric or outside [0, 625].
    """
    if val is None:
        return None
    if isinstance(val, float):
        val = int(val)
    if isinstance(val, int):
        return val if 0 <= val <= 625 else None
    s = str(val).strip()
    s = re.sub(r'^[#*+\s]+', '', s)
    s = re.sub(r'[*+\s]+$', '', s)
    try:
        v = int(float(s))
        return v if 0 <= v <= 625 else None
    except (ValueError, TypeError):
        return None


def is_nonstd(val):
    """True when val is a non-numeric string (AQA, #+matric, etc.)."""
    if val is None or isinstance(val, (int, float)):
        return False
    return bool(re.search(r'[A-Za-z]', str(val).strip()))


def normalise_code(code):
    return str(code).strip().upper()


def normalise_name(name):
    return re.sub(r'\s+', ' ', str(name).strip())


# ─────────────────────────────────────────────────────────────────────────────
# Excel parser  (2020-2025)
# ─────────────────────────────────────────────────────────────────────────────

# Maps canonical key → list of header strings to look for (upper-cased)
_XLSX_ALIASES = {
    'category':       ['CATEGORY (I.E.ISCED DESCRIPTION)',
                       'CATEGORY (ISCED DESCRIPTION)',
                       'CATEGORY (ISCED  DESCRIPTION)'],
    'course_name':    ['COURSE TITLE', 'COURSE NAME'],
    'course_code':    ['COURSE CODE', 'COURSE CODE2'],
    'eos_points':     ['EOS POINTS', 'EOS'],
    'eos_random':     ['EOS RANDOM *', 'EOS RANDOM'],
    'eos_mid':        ['EOS MID', 'EOS MIDPOINTS', 'EOS MIDPOINT', 'EOS MID-POINT'],
    'course_level':   ['COURSE LEVEL', 'LEVEL'],
    'hei':            ['HEI'],
    'test_interview': ['TEST/INTERVIEW #', 'TEST INTERVIEW #',
                       'TEST/INTERVIEW', 'TEST INTERVIEW'],
    'avp':            ['AVP'],
    'v':              ['V'],
}


def _find_xlsx_header(ws):
    """
    Scan rows until we find the header row (contains COURSE CODE or COURSE CODE2).
    Returns (row_index_0based, col_map {canonical -> 0-based col index}).
    """
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        row_up = [str(v).strip().upper() if v is not None else '' for v in row]
        if 'COURSE CODE' in row_up or 'COURSE CODE2' in row_up:
            col_map = {}
            for canon, aliases in _XLSX_ALIASES.items():
                for alias in aliases:
                    if alias in row_up:
                        col_map[canon] = row_up.index(alias)
                        break
            return ri, col_map
    return None, {}


def parse_xlsx(year):
    path = DATA_DIR / f"CAOPointsCharts{year}.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_ri, col_map = _find_xlsx_header(ws)
    if header_ri is None:
        print(f"  ⚠  Could not find header in {path.name}", file=sys.stderr)
        return []

    missing = {'course_code', 'course_name', 'hei', 'course_level'} - col_map.keys()
    if missing:
        print(f"  ⚠  Missing columns {missing} in {path.name}", file=sys.stderr)

    rows_out = []
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri <= header_ri:
            continue
        if not any(row):
            continue

        def g(canon, default=None):
            idx = col_map.get(canon)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        code  = g('course_code')
        name  = g('course_name')
        level = g('course_level')
        hei   = g('hei')

        if not code or not name:
            continue

        code = normalise_code(code)
        name = normalise_name(name)

        # Level 8 only
        try:
            if int(str(level).split('.')[0]) != 8:
                continue
        except (ValueError, TypeError):
            continue

        institution = str(hei).strip() if hei else ''
        eos_raw     = g('eos_points')
        test_int    = g('test_interview')

        # Determine capped: Test/Interview column has '#', or points are non-numeric
        capped_by_col    = bool(test_int and '#' in str(test_int))
        capped_by_points = is_nonstd(eos_raw)
        is_capped = int(capped_by_col or capped_by_points)

        # Extract numeric points even for capped courses if possible
        points = to_int_points(eos_raw)

        rows_out.append({
            'year':                    year,
            'institution':             institution,
            'course_code':             code,
            'course_name':             name,
            'points':                  points,
            'is_capped':               is_capped,
            'is_anomaly':              int(year in ANOMALY_YEARS),
            'grade_adjustment_factor': GRADE_ADJ.get(year, 0.0),
            'match_confidence':        1.0,
            'source':                  'xlsx',
        })

    return rows_out


# ─────────────────────────────────────────────────────────────────────────────
# PDF CAO parser  (2017-2019)
# ─────────────────────────────────────────────────────────────────────────────

# Course code: 2-5 uppercase letters + 2-5 digits + optional letter/digit
_COURSE_CODE_RE = re.compile(r'^([A-Z]{2,5}\d{2,5}[A-Z]?\d?)\s+')

# End of a course line: optional # then digits then optional * then space then digits
# Captures: (eos_field, eos_digits, mid_digits)
_POINTS_TAIL_RE = re.compile(
    r'\s+((?:AQA|#?\d+[*]?))\s+(\d+)\s*$'
)

_SKIP_RE = re.compile(
    r'^(admission data|end of season|level\s+[0-9]|the details|'
    r'\* not all|\*\s*not all|# test|aqa all|avp\b|v\s*=|course code\s+inst|'
    r'random|r1\s*=|page\s+\d+|^\d+\s*$|2019 level)',
    re.IGNORECASE,
)
_HEADER_COL_RE = re.compile(
    r'course\s+code\s+institution', re.IGNORECASE
)


def _is_skip(line):
    s = line.strip()
    if not s:
        return True
    if _SKIP_RE.match(s):
        return True
    if _HEADER_COL_RE.search(s):
        return True
    return False


def parse_pdf_cao(year):
    path = DATA_DIR / f"CAOPointsCharts{year}.pdf"
    rows_out = []
    current_institution = ''

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.splitlines():
                line = line.rstrip()
                if _is_skip(line):
                    continue

                code_m = _COURSE_CODE_RE.match(line)
                if not code_m:
                    # Potential institution name change
                    stripped = line.strip()
                    if stripped and not re.match(r'^\d', stripped):
                        current_institution = stripped
                    continue

                code    = normalise_code(code_m.group(1))
                rest    = line[code_m.end():]
                tail_m  = _POINTS_TAIL_RE.search(rest)

                if not tail_m:
                    # Line doesn't end with points — skip
                    continue

                eos_field = tail_m.group(1).strip()   # e.g. '290', '#615', 'AQA'
                # mid_field = tail_m.group(2)          # midpoint — not stored
                name = rest[:tail_m.start()].strip()

                capped    = eos_field.startswith('#') or eos_field.upper() == 'AQA'
                num_m     = re.search(r'\d+', eos_field)
                points    = None
                if num_m:
                    v = int(num_m.group())
                    points = v if 0 <= v <= 625 else None

                rows_out.append({
                    'year':                    year,
                    'institution':             current_institution,
                    'course_code':             code,
                    'course_name':             name,
                    'points':                  points,
                    'is_capped':               int(capped),
                    'is_anomaly':              int(year in ANOMALY_YEARS),
                    'grade_adjustment_factor': GRADE_ADJ.get(year, 0.0),
                    'match_confidence':        1.0,
                    'source':                  'pdf',
                })

    return rows_out


# ─────────────────────────────────────────────────────────────────────────────
# LC Points PDF parser
# ─────────────────────────────────────────────────────────────────────────────

# Detailed band:  "Candidates scoring between 610 - 624 points 772 1.3% 1509 2.5%"
_RE_BETWEEN = re.compile(
    r'Candidates scoring between\s+(\d+)\s*[-\u2013]\s*(\d+)\s+points\s+'
    r'([\d,]+)\s+[\d.]+%\s+([\d,]+)\s+([\d.]+)%'
)
# Exact 625:      "Candidates scoring 625 points 157 0.3% 157 0.3%"
_RE_EXACT = re.compile(
    r'Candidates scoring 625 points\s+([\d,]+)\s+[\d.]+%\s+([\d,]+)\s+([\d.]+)%'
)
# Less than 100:  "Candidates scoring less than 100 3,760 6.7% 55,770 100.0%"
_RE_LESS = re.compile(
    r'Candidates scoring less than 100\s+([\d,]+)\s+[\d.]+%\s+([\d,]+)\s+([\d.]+)%'
)


def _int_str(s):
    return int(str(s).replace(',', '').strip())


def parse_lc_pdf(year):
    suffix = str(year)[-2:]
    path   = DATA_DIR / f"lc{suffix}pts.pdf"
    rows_out = []
    seen = set()  # deduplicate by (lower, upper)

    with pdfplumber.open(path) as pdf:
        full_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)

    # Exact 625 — use search() so we only take the first match
    m = _RE_EXACT.search(full_text)
    if m:
        count   = _int_str(m.group(1))
        cum_pct = float(m.group(3))
        key = (625, 625)
        if key not in seen:
            seen.add(key)
            rows_out.append({
                'year': year, 'points_band_lower': 625, 'points_band_upper': 625,
                'student_count': count, 'cumulative_pct': cum_pct,
            })

    # Detailed bands (10-point ranges)
    for m in _RE_BETWEEN.finditer(full_text):
        lower   = int(m.group(1))
        upper   = int(m.group(2))
        count   = _int_str(m.group(3))
        cum_pct = float(m.group(5))
        key = (lower, upper)
        if key not in seen:
            seen.add(key)
            rows_out.append({
                'year': year, 'points_band_lower': lower, 'points_band_upper': upper,
                'student_count': count, 'cumulative_pct': cum_pct,
            })

    # Less than 100
    m = _RE_LESS.search(full_text)
    if m:
        count   = _int_str(m.group(1))
        cum_pct = float(m.group(3))
        key = (0, 99)
        if key not in seen:
            seen.add(key)
            rows_out.append({
                'year': year, 'points_band_lower': 0, 'points_band_upper': 99,
                'student_count': count, 'cumulative_pct': cum_pct,
            })

    return rows_out


# ─────────────────────────────────────────────────────────────────────────────
# Fuzzy matching pass (institution + name for unmatched residuals)
# ─────────────────────────────────────────────────────────────────────────────

def apply_fuzzy_matching(conn):
    """
    For course codes that only appear in a subset of years, try to find
    equivalent entries via fuzzy match on (institution, course_name).
    Updates match_confidence on matched rows.
    This is a best-effort pass; exact code matches are untouched.
    """
    import pandas as pd

    df = pd.read_sql_query(
        "SELECT id, year, institution, course_code, course_name FROM course_points",
        conn
    )

    # Build a lookup: for each (year) set of codes, find codes that have <3 years
    code_years = df.groupby('course_code')['year'].nunique()
    sparse_codes = code_years[code_years < 3].index.tolist()

    if not sparse_codes:
        return

    # For each sparse code, try to match against dense codes in nearby years
    dense_df = df[~df['course_code'].isin(sparse_codes)]
    sparse_df = df[df['course_code'].isin(sparse_codes)]

    updates = []
    for _, srow in sparse_df.iterrows():
        candidates = dense_df[
            (dense_df['year'] != srow['year']) &
            (dense_df['institution'].str.lower() == srow['institution'].lower())
        ]
        if candidates.empty:
            continue
        best_score = 0
        best_code  = None
        query = srow['course_name'].lower()
        for _, crow in candidates.iterrows():
            score = fuzz.token_sort_ratio(query, crow['course_name'].lower())
            if score > best_score:
                best_score = score
                best_code  = crow['course_code']
        # Only apply if high confidence and codes look related
        if best_score >= 85 and best_code and best_code != srow['course_code']:
            updates.append((best_code, best_score / 100.0, int(srow['id'])))

    if updates:
        conn.executemany(
            "UPDATE course_points SET course_code=?, match_confidence=? WHERE id=?",
            updates
        )
        conn.commit()
        print(f"  Fuzzy-matched {len(updates)} residual rows.")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def ingest(verbose=True):
    def log(msg):
        if verbose:
            print(msg)

    log("Creating database ...")
    conn = create_db()

    # ── CAO course points ──
    course_rows = []

    for year in [2017, 2018, 2019]:
        rows = parse_pdf_cao(year)
        log(f"  CAOPointsCharts{year}.pdf -> {len(rows)} Level-8 courses")
        course_rows.extend(rows)

    for year in [2020, 2021, 2022, 2023, 2024, 2025]:
        rows = parse_xlsx(year)
        log(f"  CAOPointsCharts{year}.xlsx -> {len(rows)} Level-8 courses")
        course_rows.extend(rows)

    conn.executemany(
        """INSERT INTO course_points
           (year,institution,course_code,course_name,points,is_capped,is_anomaly,
            grade_adjustment_factor,match_confidence,source)
           VALUES (:year,:institution,:course_code,:course_name,:points,:is_capped,:is_anomaly,
                   :grade_adjustment_factor,:match_confidence,:source)""",
        course_rows,
    )
    conn.commit()
    log(f"  -> Inserted {len(course_rows)} course_point rows total.")

    log("  Running fuzzy-match pass ...")
    apply_fuzzy_matching(conn)

    # ── LC cohort stats ──
    cohort_rows = []
    for year in range(2017, 2026):
        rows = parse_lc_pdf(year)
        log(f"  lc{str(year)[-2:]}pts.pdf -> {len(rows)} bands")
        cohort_rows.extend(rows)

    conn.executemany(
        """INSERT INTO cohort_stats
           (year,points_band_lower,points_band_upper,student_count,cumulative_pct)
           VALUES (:year,:points_band_lower,:points_band_upper,:student_count,:cumulative_pct)""",
        cohort_rows,
    )
    conn.commit()
    log(f"  -> Inserted {len(cohort_rows)} cohort_stat rows total.")

    conn.close()
    log("\nIngestion complete.")


if __name__ == "__main__":
    ingest()
